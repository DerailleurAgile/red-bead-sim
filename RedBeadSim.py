# A simulation of Dr. W.E. Deming's Red Bead Experiment using random number generation to draw sample lots.
# by Christopher R Chapman Jan 2022, updated July 2024
#
# I've written this simulation to test Dr. Deming's assertion that if the Red Bead Experiment were run
# using samples drawn using random numbers instead of mechanical sampling, the cumulative average of red 
# beads drawn over time would mirror the distribution of red to white beads (20%) according to the lot size 
# of the paddle used to draw samples (20% of 50, or 10)
#(See: Out of the Crisis, pp. 351-353)
#
# Arguments:
#--experimentCycles <int>: How many Red Bead Experiments to run; defaults to 10
#--cumulativeAvgCycles <int>: How many master cycles to run experiments within; defaults to 1
#--customSampleMethod: Flag to use my own method to draw samples; omit to default to Random.Sample()
#--baselineSamplePeriod <int>: How many samples to calculate avg and limits against; defaults to ALL (-1) if omitted
#--paddleLotSize <int>: How many beads to sample per turn; defaults to 50 for the classic experiment
#--showSigmaUnitHighlights <int>: Show visual aids for 1, 2, or 3 sigma units of dispersion around the mean; defaults to 0
#--showDistribution: Flag to show red beads in another browser tab as a distribution histogram with process limits 
#--exportToExcel: Flag to export results to an Excel workbook in the execution folder
#--showDegreesOfFreedom: Flag to show graph of uncertainty in limits

import argparse
import os
import random as rnd
from datetime import datetime
from random import shuffle

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Deming used two different configurations of the experiment: 3000 White to 750 Red, as described in Out of the Crisis,
# and 3200 White to 800 Red, as described in The New Economics
RED_BEADS_IN_BUCKET = 800
WHITE_BEADS_IN_BUCKET = 3200
BEAD_BUCKET_ARRAY = [0] * (RED_BEADS_IN_BUCKET + WHITE_BEADS_IN_BUCKET)

# Define simple flags to identify the type of bead in the bucket or paddle
RED_BEAD = 1
WHITE_BEAD = 0

# Define simple flags for determining whether to calculate 3-sigma units above or below the mean
UPPER_PROC_LIMIT = 1
LOWER_PROC_LIMIT = 0

# Define label text to identify the default sampling method used to populate the paddle
SAMPLE_METHOD = "Random.Sample()"

# Why 24? In the original Red Bead Experiment, six "willing workers" are employed for three days to pull
# one sample each. Half are fired for poor performance at the end of the third day, leaving the remainder
# to carry on with double-shifts. We dispense with the rating and ranking part of the exercise.
RED_BEAD_EXPERIMENT_LOTS = 24

# Flag to indicate calculating limits against the entire set of data points
BASELINE_PERIOD_ALL = -1

LABEL_ALIGN_CENTRE = "centre"
LABEL_ALIGN_RIGHT = "right"
LIGHT_GREEN_FILL = "90EE90"
LIGHT_ORANGE_FILL = "FFA07A"
LIGHT_RED_FILL = "FFCCCC"

def main():
    """Main function to run the Red Bead Experiment simulation."""
    args = parse_arguments()

    if args.showDegreesOfFreedom:
        plot_degrees_of_freedom(100)
    else:
        redbead_array = []
        cum_avg_log = []
        total_red_beads = 0
        cum_avg_total = 0 
        sample_count = args.experimentCycles * RED_BEAD_EXPERIMENT_LOTS

        initialize_bead_bucket(BEAD_BUCKET_ARRAY)
        for _ in range(0,args.cumulativeAvgCycles):
            redbead_array, total_red_beads = run_experiment_cycle(BEAD_BUCKET_ARRAY, sample_count, args.paddleLotSize, args.customSampleMethod)

            # Calculate cumulative average for the "day"    
            cum_avg_log.append(round(np.mean(redbead_array[:args.baselineSamplePeriod]),2))
            cum_avg_total = cum_avg_total + (total_red_beads / sample_count)
            total_red_beads = 0
        
        print_results(args, cum_avg_log, cum_avg_total, sample_count)
        
        if args.showDistribution:
            plot_results(redbead_array, args)
            plot_distribution(redbead_array, args)
        else:
            plot_results(redbead_array,args)

# As you'd expect...
def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Red Bead Experiment Simulation")
    parser.add_argument('--experimentCycles', type=int, default=10, help='How many experiments should we run? (default: 10)')
    parser.add_argument('--cumulativeAvgCycles', type=int, default=1, help='How many cycles to calculate the cumulative average against? (default: 1)')
    parser.add_argument('--customSampleMethod', action="store_true", help='Use custom sampling method to select beads. (default: True)')
    parser.add_argument('--baselineSamplePeriod', type=int, default=-1, help='How many data points to include for calculating limits? (default: -1 == ALL)')
    parser.add_argument('--paddleLotSize', type=int, default=50, help='How many beads should we sample per turn? (default: 50)')
    parser.add_argument('--showDistribution', action="store_true", help='Show the results as a distribution histogram. (default: False')
    parser.add_argument('--showDegreesOfFreedom', action="store_true", help='Show a plot of the degrees of freedom (uncertainty) for calculating limits. (default: False')
    parser.add_argument('--exportToExcel', action="store_true", help="Export simulation data to an Excel worksheet. (default: False)")
    parser.add_argument('--showSigmaUnitHighlights', type=int, default=0, help="Show transparent sigma unit highlight boxes. (1,2,3)")
    return parser.parse_args()

# In Out of the Crisis, Deming contends that the only way to have truly random samples drawn from
# the bucket is to number all the beads and select them at random using a corresponding table of
# random numbers. This method orders the beads in a virtual bucket with 0-3199 being white (0), and 
# 3200-3999 red (1). We'll use the ordinal index for each bead for random lookups after randomly
# shuffling the array.
def initialize_bead_bucket(bead_bucket_array):
    """Initialize the bead bucket with red and white beads."""
    #populate_beads_ordered(bead_bucket_array)
    populate_beads_random(bead_bucket_array)
    shuffle(bead_bucket_array)
    return bead_bucket_array

# The simulation engine...
def run_experiment_cycle(bead_bucket, sample_count, paddle_lot_size, customSampleMethod):
    """Run a single cycle of the Red Bead Experiment."""
    log = []
    total_red_beads = 0
    for _ in range(sample_count):
        if customSampleMethod:
            red_beads_pulled = pull_sample_from_bucket(bead_bucket, paddle_lot_size).count(RED_BEAD)
        else:
            red_beads_pulled = rnd.sample(bead_bucket, paddle_lot_size).count(RED_BEAD)
        
        log.append(red_beads_pulled)
        total_red_beads += red_beads_pulled
        shuffle(bead_bucket)
        shuffle(bead_bucket)
    
    return log, total_red_beads

# For confirming results...
def print_results(args, cum_avg_log, cum_avg_total, sample_count):
    """Print the results of the simulation to console."""
    baseline_sample_count = sample_count if args.baselineSamplePeriod == BASELINE_PERIOD_ALL else args.baselineSamplePeriod
    
    print("\n")
    print(f"Total Beads: {len(BEAD_BUCKET_ARRAY)}")
    print(f"Cumulative Average Cycles: {args.cumulativeAvgCycles}")
    print(f"Red Bead Experiment Cycles: {args.experimentCycles}")
    print(f"Paddle Lot Size: {args.paddleLotSize}")
    print(f"Samples Withdrawn per Experiment Cycle: {RED_BEAD_EXPERIMENT_LOTS}")
    print(f"Total Randomly-Drawn Sample Lots: {sample_count * args.cumulativeAvgCycles}")
    print(f"Baseline Sample Period: {baseline_sample_count}")
    print(f"Cumulative Average for Baseline Sample Period: {cum_avg_log}")
    print(f"Overall Cumulative Average: {round(cum_avg_total / args.cumulativeAvgCycles, 2)}\n\n")

# Guess what this does?
def plot_results(redbead_array,args):
    """Plots the XmR, shows sigma unit highlights, and exports to Excel if needed"""
    mean_array = get_mean_array(redbead_array,args.baselineSamplePeriod)
    upl_array = get_limits_array(redbead_array, mean_array[0], UPPER_PROC_LIMIT, 3, args)
    lpl_array = get_limits_array(redbead_array, mean_array[0], LOWER_PROC_LIMIT, 3, args)
    mR_array = get_moving_range_array(redbead_array)
    mR_UPL = get_moving_range_limits_array(mR_array)
    
    fig1 = plot_xmr_chart(redbead_array, mean_array, upl_array, lpl_array, mR_array, mR_UPL, args)

    # Draw highlight boxes for sigma unit ranges, if supplied
    if args.showSigmaUnitHighlights > 0:
        unit = 1
        while unit <= args.showSigmaUnitHighlights:
            sigma_upl_array = get_limits_array(redbead_array, mean_array[0], UPPER_PROC_LIMIT, unit, args)
            sigma_lpl_array = get_limits_array(redbead_array, mean_array[0], LOWER_PROC_LIMIT, unit, args)

            percent_within_limits = percent_red_beads_in_range(redbead_array, sigma_lpl_array[0], sigma_upl_array[0])

            # fig, upl_array, lpl_array, highlight_points, fill_color, box_label:""
            draw_highlight_box(fig1, sigma_upl_array, sigma_lpl_array, len(redbead_array), "Magenta", 0.1, f"{unit}σ<br>{percent_within_limits}%", LABEL_ALIGN_RIGHT)
            unit+=1

    fig1.show()

    if args.exportToExcel:
        export_to_excel(redbead_array, mean_array, mR_array, upl_array, lpl_array, mR_UPL)

def percent_red_beads_in_range(arr, min, max):
    """Return the percentage of red beads in an array that fall between two points."""
    arr = np.array(arr)
    count = np.sum((arr >= min) & (arr <= max))
    percent_red = round((count / len(arr)) * 100, 2)

    return percent_red

# UPDATED!
def export_to_excel(redbead_array, mean_array, mR_array, upl_array, lpl_array, mr_upl_array):
    """Export the simulation data to an Excel worksheet."""
    moving_range_array = [None] + mR_array

    df_combined = pd.DataFrame({
        'red beads': redbead_array,
        'avg': mean_array,
        'mR': moving_range_array,
        'upl': upl_array,
        'lpl': lpl_array,
        'mR-upl': mr_upl_array
    })

    time_stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    export_filename = "redbeadsim-" + time_stamp + ".xlsx"
    df_combined.to_excel(export_filename, index=False)
    export_filename = os.getcwd() + "/" + export_filename
    print(f"Simulation data exported to: {export_filename}")

    # Highlight Rule 1 & 2 data points if they exist
    rule1_above_indices, rule1_below_indices = get_rule1_indices(redbead_array, upl_array[0], lpl_array[0])
    rule2_above_indices, rule2_below_indices = get_rule2_indices(redbead_array, mean_array[0])
    
    if rule1_above_indices or rule1_below_indices or rule2_above_indices or rule2_below_indices:
        workbook = load_workbook(export_filename)
        worksheet = workbook.active

        if rule2_above_indices:
            highlight_worksheet_cells(worksheet, rule2_above_indices, LIGHT_ORANGE_FILL, 1)
        
        if rule2_below_indices:
            highlight_worksheet_cells(worksheet, rule2_below_indices, LIGHT_ORANGE_FILL, 1)
        
        if rule1_above_indices:
            highlight_worksheet_cells(worksheet, rule1_above_indices, LIGHT_RED_FILL, 1)
        
        if rule1_below_indices:
            highlight_worksheet_cells(worksheet, rule1_below_indices, LIGHT_RED_FILL, 1)

        print(f"Worksheet includes Rule 1 or Rule 2 highlights.")
        workbook.save(export_filename)

# NEW! In an effort to follow DRY principles...
def highlight_worksheet_cells(worksheet, indices, cell_color, column):
    """Helper method to highlight cells in the worksheet."""
    fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
    for row_idx in indices:
        cell = worksheet.cell(row=row_idx + 2, column=column)  # +2 to skip header and 1-based indexing
        cell.fill = fill

# Return an array containing the upper and lower process limits as repeating
# values for plotting on a chart
def get_limits_array(redbead_array, mean, limit_type, sigma_units, args):
    """Return an array of process limits for plotting."""
    mR_BAR = get_mr_bar(redbead_array,args.baselineSamplePeriod)
    
    if sigma_units < 0 or sigma_units > 3:
        # throw error here
        raise ValueError("Sigma units must be between 1 and 3 for calculating process limits.")

    if limit_type == UPPER_PROC_LIMIT:
        proc_limit = round(mean + sigma_units * mR_BAR / 1.128, 1)
    else:
        proc_limit = round(mean - sigma_units * mR_BAR / 1.128, 1)
        proc_limit = max(proc_limit, 0)  # Ensure proc_limit is not negative
    
    return [proc_limit] * len(redbead_array)
 
# Calculate the average moving range (mR-bar) of a set of values in an array
# and return as an integer
def get_mr_bar(redbead_array,baseline_sample_period):
    """Calculate the average moving range (mR-bar) of a set of values."""
    moving_range_array = get_moving_range_array(redbead_array)
    sample_count = len(moving_range_array) if baseline_sample_period == BASELINE_PERIOD_ALL else baseline_sample_period - 1
    mR_BAR = np.mean(moving_range_array[:sample_count])

    return round(mR_BAR, 2)

# Returns an array of repeating values representing the mean of a given array.
# We need to do this to draw the mean and process limits on the chart.
def get_mean_array(redbead_array,baseline_sample_period):
    """Return an array of the mean value repeated for plotting as a line in the XmR chart"""
    sample_count = len(redbead_array) if baseline_sample_period == BASELINE_PERIOD_ALL else baseline_sample_period
    mean = round(np.mean(redbead_array[:sample_count]), 2)
    return [mean] * len(redbead_array)

# Calculate the moving ranges and return them as an array of n-1 deltas
def get_moving_range_array(redbead_array):
    """Calculate the moving ranges and return them as an array of n-1 deltas."""
    movingRangeArray = []
    for x in range(len(redbead_array)-1):
        mR = abs(redbead_array[x+1] - redbead_array[x])
        movingRangeArray.append(mR)
    
    return movingRangeArray

# Given a array containing moving range deltas, estimate the upper limit
# as 3.268 *  the mean, and return as an array of repeating values. 
# This will be used to draw the upper limit line on an mR chart.
def get_moving_range_limits_array(moving_range_array):
    """Estimate the upper limit for moving ranges and return as an array."""
    if not moving_range_array:
        return []

    mR_BAR = sum(moving_range_array) / len(moving_range_array)
    mR_UPL = mR_BAR * 3.268

    # Why +1? Because the moving range is always shown offset by one
    return [mR_UPL] * (len(moving_range_array) + 1)

# Plots the two complementary charts in a PBC
# Yes, there are a lot of arguments to pass in...
def plot_xmr_chart(redbead_array, mean_array, upl_array, lpl_array, moving_range_array, mR_UPL, args):
    """Plot the Red Bead Experiment X-mR chart"""
     # Number of points to highlight for the baseline period
    baseline_highlight_points = args.baselineSamplePeriod

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.1,
        row_heights=[0.8, 0.2] # Symmetry! 80/20 split
    )
    
    # Add the red bead plot to the top subplot
    fig.add_trace(
        go.Scatter(
            x=list(range(len(redbead_array))),
            y=redbead_array,
            name='Individuals (X)',
            line=dict(color='royalblue', width=2),
            mode='lines+markers'
        ),
        row=1, col=1
    )

    fig.add_trace(
        go.Scatter(
            x=list(range(len(mean_array))),
            y=mean_array,
            name='Mean',
            line=dict(color='green', width=2)
        ),
        row=1, col=1
    )

    fig.add_trace(
        go.Scatter(
            x=list(range(len(upl_array))),
            y=upl_array,
            name='UPL',
            line=dict(color='red', width=2)
        ),
        row=1, col=1
    )

    fig.add_trace(
        go.Scatter(
            x=list(range(len(lpl_array))),
            y=lpl_array,
            name='LPL',
            line=dict(color='red', width=2)
        ),
        row=1, col=1
    )

    # Highlight the first 'x' samples if set
    if args.baselineSamplePeriod != BASELINE_PERIOD_ALL:
        draw_highlight_box(fig, upl_array, lpl_array, baseline_highlight_points, "LightSkyBlue", 0.3, "Baseline Period")

    # Highlight any Rule 2 signals in the X-Chart
    plot_rule_2_signals(fig, redbead_array, mean_array)

    # Add the moving range plot to the bottom subplot with an offset of one
    # Why? Because moving ranges are pairwise deltas between successive data points
    fig.add_trace(
        go.Scatter(
            x=list(range(1, len(moving_range_array) + 1)),
            y=moving_range_array,
            name='Moving Range (mR)',
            line=dict(color='royalblue', width=2),
            mode='lines+markers'
        ),
        row=2, col=1
    )

    # Add the upper process limit for the moving range -- I call it mR_UPL or "natural process limit"
    # to distinguish it from the other limits.
    fig.add_trace(
        go.Scatter(
            x=list(range(len(mR_UPL))),
            y=mR_UPL,
            name='mR-UPL',
            line=dict(color='red', width=2)
        ),
        row=2, col=1
    )

    # Update the top chart title
    chart_title_text = "Red Bead Experiment Process Behaviour Chart"
    chart_title = get_chart_header(args, redbead_array, mean_array, upl_array, lpl_array, chart_title_text)

    fig.update_layout(
        title=dict(text=chart_title, xanchor="center", x=0.5)
    )

    fig.update_xaxes(title_text="Samples", row=2, col=1)
    fig.update_yaxes(title_text="<span style='font-size:12px'><b>Individuals (X)</b></span>", row=1, col=1)
    fig.update_yaxes(title_text="<span style='font-size:12px'><b>Moving Range<br>(mR)</b></span>", row=2, col=1)

    return fig

# NEW!
def plot_rule_2_signals(fig, redbead_array, mean_array):
    """Highlight Rule 2 signals in the X-chart."""
    rule2_above_indices, rule2_below_indices = get_rule2_indices(redbead_array, mean_array[0])
    # Extract the values for above and below mean runs
    above_values = [redbead_array[i] for i in rule2_above_indices]
    below_values = [redbead_array[i] for i in rule2_below_indices]

    # Highlight above mean runs
    if rule2_above_indices:
        fig.add_trace(go.Scatter(x=rule2_above_indices, y=above_values, mode='markers',
                                marker=dict(color='orange', size=7), name='Rule 2: Above'), row=1, col=1)

    # Highlight below mean runs
    if rule2_below_indices:
        fig.add_trace(go.Scatter(x=rule2_below_indices, y=below_values, mode='markers', 
                                marker=dict(color='orange', size=7), name='Rule 2: Below'), row=1, col=1)

# NEW! 
# This is a refactored method I'm not entirely happy with, but it works for now
def draw_highlight_box(fig, upl_array, lpl_array, highlight_points, fill_color, opacity_value, box_label="", align=LABEL_ALIGN_CENTRE):
    """Draw a transparent highlight box on the plot to contain a range of data points."""
    fig.add_shape(
        type="rect",
        x0=0, y0=min(min(upl_array), min(lpl_array)),
        x1=highlight_points - 1, y1=max(max(upl_array), max(lpl_array)),
        fillcolor=fill_color, opacity=opacity_value, line_width=0,
        row=1, col=1
    )

    x_align = highlight_points
    if align == LABEL_ALIGN_CENTRE:
        x_align = highlight_points / 2
        x_anchor = "center"
    elif align == LABEL_ALIGN_RIGHT:
        x_align = highlight_points + 1
        x_anchor = "left"

    fig.add_annotation(
        x=x_align,  # Positioning the text in the middle of the rectangle
        y=max(upl_array),
        text="<b>" + box_label + "</b>",
        showarrow=False,
        font=dict(size=12, color="black"),
        align="center",
        xanchor=x_anchor,
        yanchor="top",
        opacity=0.7,
        row=1, col=1
    )

    fig.update_layout()

# NEW! 
def plot_distribution(redbead_array, args):
    """Plot the distribution of red beads as a histogram"""
    fig = px.histogram(redbead_array, nbins=22, title='Red Bead Distribution')
    
    # Yes, there is a more elegant way to do this...
    mean_array = get_mean_array(redbead_array, args.baselineSamplePeriod)
    upl_array = get_limits_array(redbead_array, UPPER_PROC_LIMIT, 3, args)
    lpl_array = get_limits_array(redbead_array, LOWER_PROC_LIMIT, 3, args)

    # Add vertical lines for mean, UPL, and LPL
    fig.add_shape(
        type="line",
        x0=mean_array[0], y0=0, x1=mean_array[0], y1=1,
        xref='x', yref='paper',
        line=dict(color="Green", width=2, dash="dash"),
        name='Mean'
    )
    fig.add_shape(
        type="line",
        x0=upl_array[0], y0=0, x1=upl_array[0], y1=1,
        xref='x', yref='paper',
        line=dict(color="Red", width=2, dash="dash"),
        name='UPL'
    )
    fig.add_shape(
        type="line",
        x0=lpl_array[0], y0=0, x1=lpl_array[0], y1=1,
        xref='x', yref='paper',
        line=dict(color="Red", width=2, dash="dash"),
        name='LPL'
    )

    fig.update_layout(
        xaxis_title='Red Beads',
        yaxis_title='Count',
        bargap=0.1,
        title={
            'text': "Red Bead Distribution Chart",
            'y':0.9,
            'x':0.5,
            'xanchor': 'center',
            'yanchor': 'top'
        }
    )

    # Update the top chart title
    chart_title_text = "Red Bead Distribution Chart"
    chart_title = get_chart_header(args, redbead_array, mean_array, upl_array, lpl_array, chart_title_text)
    fig.update_layout(
        title=dict(text=chart_title, x=0.5)
    )

    fig.show()

# NEW!
def plot_degrees_of_freedom(max_points):
    """Plot the degrees of freedom or uncertainty in process limits from 1 to max_points"""
    points = np.arange(1, max_points + 1)
    degrees_of_freedom = calculate_degrees_of_freedom(points)
    
    df = pd.DataFrame({
        'Data Points': points,
        'Degrees of Freedom (Uncertainty)': degrees_of_freedom * 100
    })
    
    fig = px.line(df, x='Data Points', y='Degrees of Freedom (Uncertainty)',
                  title='Degrees of Freedom in Computed Limits for a Process Behaviour Chart',
                  labels={'Data Points': 'Number of Data Points', 'Degrees of Freedom (Uncertainty)': 'Degrees of Freedom (Uncertainty)'})

    fig.update_yaxes(tickformat=".0f", ticksuffix="%")
    fig.show()

# NEW!
def calculate_degrees_of_freedom(num_points):
    """Calculate the degrees of freedom for a given number of data points."""
    return 1 / np.sqrt(2 * num_points)

#NEW!
def get_rule1_indices(redbead_array, upl, lpl):
    """Find Rule 1 signals in an array of red bead values."""
    rule1_above_indices = []
    rule1_below_indices = []

    for i, value in enumerate(redbead_array):
        if value > upl:
            rule1_above_indices.append(i)
        elif value < lpl:
            rule1_below_indices.append(i)

    return rule1_above_indices, rule1_below_indices

# NEW!
def get_rule2_indices(redbead_array, mean_value, min_run_length=8):
    """Find Rule 2 signals in an array of red bead values."""
    def append_rule2_indices(rule2_signal_run, rule2_signal_run_list):
        if len(rule2_signal_run) >= min_run_length:
            rule2_signal_run_list.append(rule2_signal_run)
    
    rule2_above_indices = []
    rule2_below_indices = []
    current_above_run = []
    current_below_run = []
    
    for i, value in enumerate(redbead_array):
        if value > mean_value:
            if current_below_run:
                append_rule2_indices(current_below_run, rule2_below_indices)
                current_below_run = []
            current_above_run.append(i)
        else:
            if current_above_run:
                append_rule2_indices(current_above_run, rule2_above_indices)
                current_above_run = []
            current_below_run.append(i)

    # Check for last run, if any
    append_rule2_indices(current_above_run, rule2_above_indices)
    append_rule2_indices(current_below_run, rule2_below_indices)

    # Flatten nested lists
    flattened_above_indices = [index for sublist in rule2_above_indices for index in sublist]
    flattened_below_indices = [index for sublist in rule2_below_indices for index in sublist]
    
    return flattened_above_indices, flattened_below_indices

# NEW!
def get_chart_header(args, redbead_array, mean_array, upl_array, lpl_array, chart_title_text=""):
    """Generate the chart header text."""
    baseline_sample_count = len(redbead_array) if args.baselineSamplePeriod == BASELINE_PERIOD_ALL else args.baselineSamplePeriod
    chart_title = (
        f"<span style='font-weight:bold'>{chart_title_text}</span><br>"
        f"<span style='font-size:12px'>"
        f"<b>Experiments: </b>{args.experimentCycles} "
        f"<b>Paddle Size: </b>{args.paddleLotSize} "
        f"<b>Data Points:</b> {len(redbead_array)} "
        f"<b>Method: </b>{SAMPLE_METHOD} "
        f"<b>Baseline Sample Period: </b>{baseline_sample_count} "
        f"<b>Total Red Beads: </b>{sum(redbead_array)} "
        f"<br><b>Mean:</b> {mean_array[0]} "
        f"<b>UPL:</b> {upl_array[0]} "
        f"<b>LPL:</b> {lpl_array[0]}</span>"
    )
    return chart_title

# My custom method for drawing a random sample of beads from the bucket
def pull_sample_from_bucket(bucket_array,paddle_size):
    """Custom method for drawing a random sample of beads from the bucket."""
    sample_array = []
    paddle_index_log = []
    sample_count = 0
    global SAMPLE_METHOD # this needs fixin
    SAMPLE_METHOD = "Custom"

    while sample_count < paddle_size:
        index = rnd.randint(0, len(bucket_array) - 1)
        if index not in paddle_index_log:
            paddle_index_log.append(index)
            sample_array.append(bucket_array[index])
            sample_count += 1

    return sample_array

# Randomly fills the "sample bucket" with red and white beads
def populate_beads_random(bucket_array):
    """Randomly fill the sample bucket with red and white beads."""
    red_beads_count = 0
    total_beads = len(bucket_array)
    while red_beads_count <= RED_BEADS_IN_BUCKET:
        index = rnd.randint(0,total_beads- 1)
        if(bucket_array[index] == 0):
            bucket_array[index] = RED_BEAD
            red_beads_count += 1

# Fills the sample bucket with 3,2000 white beads first, then 800 red beads
def populate_beads_ordered(bucket_array):
    """Fill the sample bucket with white beads first, then red beads."""
    for bead in range(len(bucket_array)):
        bucket_array[bead] = WHITE_BEAD if bead < WHITE_BEADS_IN_BUCKET else RED_BEAD

if __name__ == "__main__":
    main()
